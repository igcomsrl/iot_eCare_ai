from DataRetriever import DataRetriever
from datetime import datetime
import sys
from openpyxl import Workbook
import pandas as pd
import pytz
import json
import os

def main(processInstanceId: str, date1: datetime, date2: datetime):
    metrics = ['temperature', 'systolic', 'diastolic', 'pulse', 'spo2', 'weight']

    ipDB = os.getenv('INFLUX_IP_AI', 'localhost')
    portDB = os.getenv('INFLUX_PORT_AI', '8086')
    userDB = os.getenv('INFLUX_USER_AI', 'admin')
    passwordDB = os.getenv('INFLUX_PW_AI', 'G10m1R0m3')
    nameDB = os.getenv('INFLUX_DB_AI', 'giomi')

    dr = DataRetriever(metrics)
    dfs = dr.loadDataFromDB(ipDB, portDB, userDB, passwordDB, nameDB)

    wb = Workbook()
    sheets = dict()
    ws = None

    with open('{}/names.json'.format(os.path.dirname(os.path.abspath(__file__))), encoding='latin1') as f:
        jsondata = json.load(f)
        for index, df in enumerate(dfs):
            toIterate = None
            if processInstanceId == 'all':
                toIterate = df['processInstanceId'].unique()
            else:
                toIterate = [processInstanceId]

            for pid in toIterate:
                if pid not in sheets:
                    if pid not in jsondata:
                        continue
                    sheets[pid] = wb.create_sheet(jsondata[pid]['firstname'] + jsondata[pid]['surname'])

                ws = sheets[pid]

                ws.cell(index*3 + 2, 1, metrics[index])
                df['time'] = pd.to_datetime(df['time'])
                minidf = df[df['processInstanceId'] == pid]
                minidf = minidf[(minidf['time'] >= date1) & (minidf['time'] <= date2)]
                # print(minidf)
                minidf.reset_index(inplace=True)
                # print(len(minidf))
                for row, data in minidf.iterrows():
                    ws.cell(index * 3 + 1, row + 2, datetime.strftime(data['time'], '%d-%m-%Y %H:%M:%S'))
                    ws.cell(index * 3 + 2, row + 2, data[metrics[index]])

        for _,value in sheets.items():
            for col in value.iter_cols():
                value.column_dimensions[col[0].column_letter].width = 23

        wb.save(os.path.dirname(os.path.abspath(__file__)) + '/out.xlsx')

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print("Usage: python3 main.py <processInstanceId> <date1> <date2>\nDates should be given in the following format: dd-mm-aaaa")
        exit()
    
    processInstanceId = sys.argv[1]
    try:
        date1 = datetime.strptime(sys.argv[2], '%d-%m-%Y')
        date2 = datetime.strptime(sys.argv[3], '%d-%m-%Y')
    except ValueError:
        print("Dates should be given in the following format: dd-mm-aaaa")
        exit(1)

    timezone = pytz.timezone('UTC')
    date1 = timezone.localize(date1)
    date2 = timezone.localize(date2)

    main(processInstanceId, date1, date2)
